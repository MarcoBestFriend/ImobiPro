"""
================================================================================
IMOBIPRO - SISTEMA DE BACKUP
================================================================================
Autor: Sistema ImobiPro
Data: Janeiro 2026
Descrição: Sistema completo de backup e restauração do banco de dados.
           Permite backup do SQLite e exportação para Excel.
================================================================================
"""

import os
import shutil
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from database.db_manager import DatabaseManager

class SistemaBackup:
    """
    Classe para gerenciar backups do sistema ImobiPro.
    """
    
    def __init__(self, db_manager: DatabaseManager = None):
        """
        Inicializa o sistema de backup.
        
        Args:
            db_manager: Instância do DatabaseManager (opcional)
        """
        self.db = db_manager or DatabaseManager()
        self.dir_backups = 'backups'
        
        # Criar diretório de backups se não existir
        os.makedirs(self.dir_backups, exist_ok=True)
    
    def gerar_nome_arquivo(self, tipo: str = 'db', extensao: str = 'db') -> str:
        """
        Gera nome de arquivo para backup com timestamp.
        
        Args:
            tipo (str): Tipo do backup (db, excel)
            extensao (str): Extensão do arquivo
        
        Returns:
            str: Nome do arquivo de backup
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"imobipro_backup_{tipo}_{timestamp}.{extensao}"
    
    def backup_sqlite(self) -> tuple[bool, str]:
        """
        Cria backup do banco de dados SQLite.
        
        Returns:
            tuple: (sucesso, caminho_arquivo ou mensagem_erro)
        """
        print("\n" + "="*70)
        print("CRIANDO BACKUP DO BANCO DE DADOS SQLITE")
        print("="*70)
        
        try:
            # Caminho do arquivo original
            db_original = self.db.db_path
            
            if not os.path.exists(db_original):
                return False, "Banco de dados não encontrado"
            
            # Gerar nome do backup
            nome_backup = self.gerar_nome_arquivo('db', 'db')
            caminho_backup = os.path.join(self.dir_backups, nome_backup)
            
            # Copiar arquivo do banco
            shutil.copy2(db_original, caminho_backup)
            
            # Verificar se o backup foi criado
            if os.path.exists(caminho_backup):
                tamanho = os.path.getsize(caminho_backup)
                print(f"✓ Backup criado com sucesso!")
                print(f"  Arquivo: {caminho_backup}")
                print(f"  Tamanho: {tamanho:,} bytes")
                return True, caminho_backup
            else:
                return False, "Erro ao criar arquivo de backup"
                
        except Exception as e:
            return False, f"Erro ao criar backup: {str(e)}"
    
    def exportar_para_excel(self) -> tuple[bool, str]:
        """
        Exporta todos os dados do banco para um arquivo Excel.
        
        Returns:
            tuple: (sucesso, caminho_arquivo ou mensagem_erro)
        """
        print("\n" + "="*70)
        print("EXPORTANDO DADOS PARA EXCEL")
        print("="*70)
        
        try:
            # Criar workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remover sheet padrão
            
            # Definir tabelas a exportar
            tabelas = ['imoveis', 'pessoas', 'contratos', 'despesas', 'receitas']
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center")
            
            for tabela in tabelas:
                print(f"\nProcessando tabela: {tabela.upper()}")
                
                # Buscar dados
                dados = self.db.get_all(tabela)
                
                if not dados:
                    print(f"  ⚠ Tabela {tabela} está vazia")
                    continue
                
                # Criar sheet
                ws = wb.create_sheet(title=tabela.upper())
                
                # Adicionar cabeçalhos
                headers = list(dados[0].keys())
                ws.append(headers)
                
                # Estilizar cabeçalhos
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                
                # Adicionar dados
                for registro in dados:
                    row = [registro[header] for header in headers]
                    ws.append(row)
                
                # Ajustar largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                print(f"  ✓ {len(dados)} registros exportados")
            
            # Salvar arquivo
            nome_excel = self.gerar_nome_arquivo('excel', 'xlsx')
            caminho_excel = os.path.join(self.dir_backups, nome_excel)
            wb.save(caminho_excel)
            
            print(f"\n✓ Exportação concluída com sucesso!")
            print(f"  Arquivo: {caminho_excel}")
            
            return True, caminho_excel
            
        except Exception as e:
            return False, f"Erro ao exportar para Excel: {str(e)}"
    
    def restaurar_sqlite(self, caminho_backup: str) -> tuple[bool, str]:
        """
        Restaura o banco de dados a partir de um backup.
        
        Args:
            caminho_backup (str): Caminho do arquivo de backup
        
        Returns:
            tuple: (sucesso, mensagem)
        """
        print("\n" + "="*70)
        print("RESTAURANDO BANCO DE DADOS")
        print("="*70)
        
        try:
            if not os.path.exists(caminho_backup):
                return False, f"Arquivo de backup não encontrado: {caminho_backup}"
            
            # Criar backup do banco atual antes de restaurar
            print("\nCriando backup de segurança do banco atual...")
            backup_seguranca = os.path.join(
                self.dir_backups,
                f"pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            )
            
            if os.path.exists(self.db.db_path):
                shutil.copy2(self.db.db_path, backup_seguranca)
                print(f"✓ Backup de segurança criado: {backup_seguranca}")
            
            # Fechar conexão atual
            self.db.close()
            
            # Restaurar backup
            shutil.copy2(caminho_backup, self.db.db_path)
            
            # Reconectar
            self.db.connect()
            
            # Verificar integridade
            integridade_ok, erros = self.db.verificar_integridade()
            
            if integridade_ok:
                print("\n✓ Banco de dados restaurado com sucesso!")
                print("✓ Integridade verificada!")
                return True, "Restauração concluída com sucesso"
            else:
                print("\n⚠ Banco restaurado mas com problemas de integridade:")
                for erro in erros:
                    print(f"  - {erro}")
                return True, "Restaurado com avisos de integridade"
                
        except Exception as e:
            return False, f"Erro ao restaurar backup: {str(e)}"
    
    def listar_backups(self) -> list:
        """
        Lista todos os backups disponíveis.
        
        Returns:
            list: Lista de dicionários com informações dos backups
        """
        backups = []
        
        try:
            arquivos = os.listdir(self.dir_backups)
            arquivos = [f for f in arquivos if f.startswith('imobipro_backup_')]
            arquivos.sort(reverse=True)  # Mais recentes primeiro
            
            for arquivo in arquivos:
                caminho = os.path.join(self.dir_backups, arquivo)
                stat = os.stat(caminho)
                
                backups.append({
                    'nome': arquivo,
                    'caminho': caminho,
                    'tamanho': stat.st_size,
                    'data': datetime.fromtimestamp(stat.st_mtime),
                    'tipo': 'SQLite' if arquivo.endswith('.db') else 'Excel'
                })
            
        except Exception as e:
            print(f"Erro ao listar backups: {e}")
        
        return backups
    
    def limpar_backups_antigos(self, manter_ultimos: int = 10) -> int:
        """
        Remove backups antigos, mantendo apenas os N mais recentes.
        
        Args:
            manter_ultimos (int): Quantidade de backups a manter
        
        Returns:
            int: Quantidade de backups removidos
        """
        backups = self.listar_backups()
        removidos = 0
        
        if len(backups) > manter_ultimos:
            for backup in backups[manter_ultimos:]:
                try:
                    os.remove(backup['caminho'])
                    removidos += 1
                    print(f"✓ Removido: {backup['nome']}")
                except Exception as e:
                    print(f"✗ Erro ao remover {backup['nome']}: {e}")
        
        return removidos
    
    def backup_completo(self, limpar_antigos: bool = True) -> dict:
        """
        Executa backup completo (SQLite + Excel).
        
        Args:
            limpar_antigos (bool): Se deve limpar backups antigos
        
        Returns:
            dict: Resultado das operações
        """
        print("\n" + "="*70)
        print("EXECUTANDO BACKUP COMPLETO")
        print("="*70)
        print(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        
        resultado = {
            'sucesso': False,
            'backup_sqlite': None,
            'backup_excel': None,
            'erros': []
        }
        
        # Backup SQLite
        sucesso_db, msg_db = self.backup_sqlite()
        if sucesso_db:
            resultado['backup_sqlite'] = msg_db
        else:
            resultado['erros'].append(f"SQLite: {msg_db}")
        
        # Exportar para Excel
        sucesso_excel, msg_excel = self.exportar_para_excel()
        if sucesso_excel:
            resultado['backup_excel'] = msg_excel
        else:
            resultado['erros'].append(f"Excel: {msg_excel}")
        
        # Limpar backups antigos
        if limpar_antigos:
            print("\n" + "-"*70)
            print("Limpando backups antigos...")
            removidos = self.limpar_backups_antigos(manter_ultimos=10)
            if removidos > 0:
                print(f"✓ {removidos} backups antigos removidos")
        
        # Resultado final
        resultado['sucesso'] = sucesso_db and sucesso_excel
        
        if resultado['sucesso']:
            print("\n" + "="*70)
            print("✓ BACKUP COMPLETO REALIZADO COM SUCESSO!")
            print("="*70)
        else:
            print("\n" + "="*70)
            print("⚠ BACKUP CONCLUÍDO COM ERROS")
            print("="*70)
            for erro in resultado['erros']:
                print(f"  ✗ {erro}")
        
        return resultado


# ============================================================================
# FUNÇÕES DE UTILIDADE
# ============================================================================

def formatar_tamanho(bytes: int) -> str:
    """
    Formata tamanho de arquivo em formato legível.
    
    Args:
        bytes (int): Tamanho em bytes
    
    Returns:
        str: Tamanho formatado (ex: "1.5 MB")
    """
    for unidade in ['B', 'KB', 'MB', 'GB']:
        if bytes < 1024.0:
            return f"{bytes:.2f} {unidade}"
        bytes /= 1024.0
    return f"{bytes:.2f} TB"


# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    """
    Executa backup quando o script é rodado diretamente.
    """
    
    # Criar sistema de backup
    backup = SistemaBackup()
    
    # Menu simples
    print("\n" + "="*70)
    print("SISTEMA DE BACKUP IMOBIPRO")
    print("="*70)
    print("\nOpções:")
    print("1. Backup completo (SQLite + Excel)")
    print("2. Backup somente SQLite")
    print("3. Exportar somente para Excel")
    print("4. Listar backups disponíveis")
    print("5. Restaurar backup")
    print("0. Sair")
    
    try:
        opcao = input("\nEscolha uma opção: ").strip()
        
        if opcao == '1':
            backup.backup_completo()
            
        elif opcao == '2':
            sucesso, msg = backup.backup_sqlite()
            if not sucesso:
                print(f"\n✗ Erro: {msg}")
                
        elif opcao == '3':
            sucesso, msg = backup.exportar_para_excel()
            if not sucesso:
                print(f"\n✗ Erro: {msg}")
                
        elif opcao == '4':
            backups = backup.listar_backups()
            if backups:
                print(f"\n{len(backups)} backup(s) disponível(is):\n")
                for i, bkp in enumerate(backups, 1):
                    print(f"{i}. {bkp['nome']}")
                    print(f"   Tipo: {bkp['tipo']}")
                    print(f"   Data: {bkp['data'].strftime('%d/%m/%Y %H:%M:%S')}")
                    print(f"   Tamanho: {formatar_tamanho(bkp['tamanho'])}")
                    print()
            else:
                print("\nNenhum backup encontrado.")
                
        elif opcao == '5':
            backups = backup.listar_backups()
            db_backups = [b for b in backups if b['tipo'] == 'SQLite']
            
            if not db_backups:
                print("\nNenhum backup SQLite disponível.")
            else:
                print("\nBackups SQLite disponíveis:\n")
                for i, bkp in enumerate(db_backups, 1):
                    print(f"{i}. {bkp['nome']} - {bkp['data'].strftime('%d/%m/%Y %H:%M:%S')}")
                
                try:
                    escolha = int(input("\nEscolha o backup para restaurar (número): "))
                    if 1 <= escolha <= len(db_backups):
                        confirmacao = input(f"\n⚠ ATENÇÃO: O banco atual será substituído!\nConfirma a restauração? (sim/não): ")
                        if confirmacao.lower() == 'sim':
                            sucesso, msg = backup.restaurar_sqlite(db_backups[escolha-1]['caminho'])
                            if not sucesso:
                                print(f"\n✗ Erro: {msg}")
                        else:
                            print("\nRestauração cancelada.")
                    else:
                        print("\nOpção inválida.")
                except ValueError:
                    print("\nEntrada inválida.")
        
        elif opcao == '0':
            print("\nSaindo...")
        else:
            print("\nOpção inválida.")
            
    except KeyboardInterrupt:
        print("\n\nOperação cancelada pelo usuário.")
    except Exception as e:
        print(f"\n✗ Erro inesperado: {e}")
