"""
================================================================================
IMOBIPRO - APLICAÇÃO WEB PRINCIPAL
================================================================================
Autor: Sistema ImobiPro
Data: Janeiro 2026
Descrição: Aplicação Flask principal com todas as rotas e funcionalidades
================================================================================
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime, date
import os
import csv
import zipfile
import io
import tempfile

from config import get_config
from database.db_manager import DatabaseManager
from utils.backup import SistemaBackup

# Criar aplicação Flask
app = Flask(__name__)
app.config.from_object(get_config('development'))

# Inicializar gerenciador de banco de dados
db = DatabaseManager()

# Inicializar sistema de backup
backup_system = SistemaBackup(db)

# ============================================================================
# CONFIGURAÇÃO DE LOGIN
# ============================================================================

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'warning'


class User(UserMixin):
    """Classe simples de usuário para Flask-Login."""
    def __init__(self, id):
        self.id = id


@login_manager.user_loader
def load_user(user_id):
    """Carrega o usuário pelo ID."""
    if user_id == 'admin':
        return User('admin')
    return None

# ============================================================================
# BACKUP AUTOMÁTICO
# ============================================================================

def executar_backup_automatico():
    """Executa backup automático do banco de dados."""
    try:
        print(f"\n[BACKUP AUTOMÁTICO] Iniciando às {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        sucesso, caminho = backup_system.backup_sqlite()
        if sucesso:
            print(f"[BACKUP AUTOMÁTICO] Concluído: {caminho}")
            # Limpar backups antigos (manter últimos 7)
            backup_system.limpar_backups_antigos(manter_ultimos=7)
        else:
            print(f"[BACKUP AUTOMÁTICO] Falha: {caminho}")
    except Exception as e:
        print(f"[BACKUP AUTOMÁTICO] Erro: {str(e)}")

def obter_ultimo_backup():
    """Retorna informações do último backup realizado."""
    try:
        backups = backup_system.listar_backups()
        db_backups = [b for b in backups if b['tipo'] == 'SQLite']
        if db_backups:
            ultimo = db_backups[0]
            return ultimo['data'].strftime('%d/%m/%Y às %H:%M')
        return None
    except:
        return None

# Configurar APScheduler para backup diário
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger

    scheduler = BackgroundScheduler()
    # Executa backup todos os dias às 02:00
    scheduler.add_job(
        executar_backup_automatico,
        CronTrigger(hour=2, minute=0),
        id='backup_diario',
        name='Backup automático diário',
        replace_existing=True
    )
    scheduler.start()
    print("✓ Backup automático configurado para 02:00 diariamente")
except ImportError:
    print("⚠ APScheduler não instalado. Backup automático desativado.")
    print("  Para ativar, execute: pip install apscheduler")
except Exception as e:
    print(f"⚠ Erro ao configurar backup automático: {e}")


# ============================================================================
# FILTROS PERSONALIZADOS PARA TEMPLATES
# ============================================================================

@app.template_filter('formatar_moeda')
def formatar_moeda(valor):
    """Formata valor como moeda brasileira."""
    if valor is None:
        return "R$ 0,00"
    return f"R$ {float(valor):,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')


@app.template_filter('formatar_data')
def formatar_data(data):
    """Formata data para formato brasileiro."""
    if not data:
        return ""
    if isinstance(data, str):
        try:
            dt = datetime.strptime(data, '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except:
            return data
    if isinstance(data, (date, datetime)):
        return data.strftime('%d/%m/%Y')
    return data


@app.template_filter('status_badge')
def status_badge(status):
    """Retorna classe CSS para badge de status."""
    badges = {
        'Ativo': 'badge-success',
        'Prorrogado': 'badge-info',
        'Encerrado': 'badge-secondary',
        'Rescindido': 'badge-danger',
        'Pendente': 'badge-warning',
        'Recebido': 'badge-success',
        'Atrasado': 'badge-danger',
        'Sim': 'badge-success',
        'Não': 'badge-secondary',
    }
    return badges.get(status, 'badge-primary')


# ============================================================================
# ROTAS - AUTENTICAÇÃO
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login."""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')

        # Verificar credenciais
        if username == app.config.get('LOGIN_USERNAME', 'admin') and \
           password == app.config.get('LOGIN_PASSWORD', 'imobipro2026'):
            user = User(username)
            login_user(user, remember=True)
            flash('Login realizado com sucesso!', 'success')

            # Redirecionar para a página original ou dashboard
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Usuário ou senha incorretos.', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Realiza logout do usuário."""
    logout_user()
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('login'))


# ============================================================================
# ROTAS - PÁGINA INICIAL E DASHBOARD
# ============================================================================

@app.route('/')
def index():
    """Página inicial - redireciona para dashboard."""
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal com estatísticas e resumos."""
    # Buscar estatísticas
    stats = db.get_estatisticas_dashboard()
    
    # Buscar contratos ativos
    contratos_ativos = db.get_contratos_ativos()
    
    # Buscar despesas pendentes (próximas a vencer)
    despesas_pendentes = db.get_despesas_pendentes()
    
    # Buscar receitas pendentes (próximas a vencer)
    receitas_pendentes = db.get_receitas_pendentes()
    
    # Imóveis disponíveis
    imoveis_disponiveis = db.get_imoveis_disponiveis()
    
    return render_template('dashboard.html',
                         stats=stats,
                         contratos_ativos=contratos_ativos[:5],  # Últimos 5
                         despesas_pendentes=despesas_pendentes[:5],
                         receitas_pendentes=receitas_pendentes[:5],
                         imoveis_disponiveis=imoveis_disponiveis[:5])


# ============================================================================
# ROTAS - IMÓVEIS
# ============================================================================

@app.route('/imoveis')
@login_required
def listar_imoveis():
    """Lista todos os imóveis."""
    # Buscar parâmetros de filtro
    filtro_ocupado = request.args.get('ocupado', '')
    busca = request.args.get('busca', '')
    
    # Buscar imóveis
    if filtro_ocupado:
        imoveis = db.get_where('imoveis', 'ocupado = ?', (filtro_ocupado,), 'endereco_completo')
    elif busca:
        imoveis = db.get_where('imoveis', 'endereco_completo LIKE ?', (f'%{busca}%',), 'endereco_completo')
    else:
        imoveis = db.get_all('imoveis', 'endereco_completo')
    
    return render_template('imoveis/listar.html', imoveis=imoveis, filtro_ocupado=filtro_ocupado, busca=busca)


@app.route('/imoveis/novo', methods=['GET', 'POST'])
@login_required
def novo_imovel():
    """Cadastra novo imóvel."""
    if request.method == 'POST':
        try:
            # Coletar dados do formulário
            dados = {
                'endereco_completo': request.form.get('endereco_completo'),
                'inscricao_imobiliaria': request.form.get('inscricao_imobiliaria'),
                'matricula': request.form.get('matricula'),
                'tipo_imovel': request.form.get('tipo_imovel'),
                'id_proprietario': request.form.get('id_proprietario') or None,
                'valor_iptu_anual': request.form.get('valor_iptu_anual') or None,
                'forma_pagamento_iptu': request.form.get('forma_pagamento_iptu', 'Anual'),
                'aluguel_pretendido': request.form.get('aluguel_pretendido') or None,
                'condominio_sugerido': request.form.get('condominio_sugerido') or None,
                'dia_venc_condominio': request.form.get('dia_venc_condominio') or None,
                'valor_mercado': request.form.get('valor_mercado') or None,
                'data_aquisicao': request.form.get('data_aquisicao') or None,
                'numero_hidrometro': request.form.get('numero_hidrometro'),
                'numero_relogio_energia': request.form.get('numero_relogio_energia'),
                'cidade': request.form.get('cidade') or 'Campo Grande',
                'estado': request.form.get('estado') or 'MS',
                'cep': request.form.get('cep'),
                'observacoes': request.form.get('observacoes'),
            }

            # Validar campo obrigatório
            if not dados['endereco_completo']:
                flash('Endereço completo é obrigatório!', 'danger')
                proprietarios = db.execute_query("SELECT id, nome_completo FROM pessoas ORDER BY nome_completo")
                return render_template('imoveis/form.html', imovel=dados, proprietarios=proprietarios, config=app.config)

            # Inserir no banco
            imovel_id = db.insert('imoveis', dados)

            if imovel_id:
                flash(f'Imóvel cadastrado com sucesso! ID: {imovel_id}', 'success')
                return redirect(url_for('listar_imoveis'))
            else:
                flash('Erro ao cadastrar imóvel.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    # Buscar proprietários para o select
    proprietarios = db.execute_query("SELECT id, nome_completo FROM pessoas ORDER BY nome_completo")
    return render_template('imoveis/form.html', imovel=None, proprietarios=proprietarios, config=app.config)


@app.route('/imoveis/<int:id>')
@login_required
def ver_imovel(id):
    """Visualiza detalhes de um imóvel."""
    imovel = db.get_by_id('imoveis', id)
    
    if not imovel:
        flash('Imóvel não encontrado.', 'danger')
        return redirect(url_for('listar_imoveis'))
    
    # Buscar contratos do imóvel
    contratos = db.get_where('contratos', 'id_imovel = ?', (id,), 'inicio_contrato DESC')
    
    # Buscar despesas do imóvel
    despesas = db.get_where('despesas', 'id_imovel = ?', (id,), 'vencimento_previsto DESC')
    
    return render_template('imoveis/ver.html', imovel=imovel, contratos=contratos, despesas=despesas)


@app.route('/imoveis/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_imovel(id):
    """Edita um imóvel existente."""
    imovel = db.get_by_id('imoveis', id)
    
    if not imovel:
        flash('Imóvel não encontrado.', 'danger')
        return redirect(url_for('listar_imoveis'))
    
    if request.method == 'POST':
        try:
            dados = {
                'endereco_completo': request.form.get('endereco_completo'),
                'inscricao_imobiliaria': request.form.get('inscricao_imobiliaria'),
                'matricula': request.form.get('matricula'),
                'tipo_imovel': request.form.get('tipo_imovel'),
                'id_proprietario': request.form.get('id_proprietario') or None,
                'valor_iptu_anual': request.form.get('valor_iptu_anual') or None,
                'forma_pagamento_iptu': request.form.get('forma_pagamento_iptu'),
                'aluguel_pretendido': request.form.get('aluguel_pretendido') or None,
                'condominio_sugerido': request.form.get('condominio_sugerido') or None,
                'dia_venc_condominio': request.form.get('dia_venc_condominio') or None,
                'valor_mercado': request.form.get('valor_mercado') or None,
                'data_aquisicao': request.form.get('data_aquisicao') or None,
                'numero_hidrometro': request.form.get('numero_hidrometro'),
                'numero_relogio_energia': request.form.get('numero_relogio_energia'),
                'cidade': request.form.get('cidade') or 'Campo Grande',
                'estado': request.form.get('estado') or 'MS',
                'cep': request.form.get('cep'),
                'observacoes': request.form.get('observacoes'),
            }

            if not dados['endereco_completo']:
                flash('Endereço completo é obrigatório!', 'danger')
                proprietarios = db.execute_query("SELECT id, nome_completo FROM pessoas ORDER BY nome_completo")
                return render_template('imoveis/form.html', imovel=dados, proprietarios=proprietarios, config=app.config)

            if db.update('imoveis', dados, 'id = ?', (id,)):
                flash('Imóvel atualizado com sucesso!', 'success')
                return redirect(url_for('ver_imovel', id=id))
            else:
                flash('Erro ao atualizar imóvel.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    # Buscar proprietários para o select
    proprietarios = db.execute_query("SELECT id, nome_completo FROM pessoas ORDER BY nome_completo")
    return render_template('imoveis/form.html', imovel=imovel, proprietarios=proprietarios, config=app.config)


@app.route('/imoveis/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_imovel(id):
    """Exclui um imóvel."""
    try:
        if db.delete('imoveis', 'id = ?', (id,)):
            flash('Imóvel excluído com sucesso!', 'success')
        else:
            flash('Erro ao excluir imóvel.', 'danger')
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
    
    return redirect(url_for('listar_imoveis'))


# ============================================================================
# ROTAS - PESSOAS
# ============================================================================

@app.route('/pessoas')
@login_required
def listar_pessoas():
    """Lista todas as pessoas."""
    situacao = request.args.get('situacao', '')
    busca = request.args.get('busca', '')
    
    if situacao:
        pessoas = db.get_where('pessoas', 'situacao = ?', (situacao,), 'nome_completo')
    elif busca:
        pessoas = db.get_where('pessoas', 'nome_completo LIKE ?', (f'%{busca}%',), 'nome_completo')
    else:
        pessoas = db.get_all('pessoas', 'nome_completo')
    
    return render_template('pessoas/listar.html', pessoas=pessoas, situacao=situacao, busca=busca)


@app.route('/pessoas/novo', methods=['GET', 'POST'])
@login_required
def nova_pessoa():
    """Cadastra nova pessoa."""
    if request.method == 'POST':
        try:
            dados = {
                'situacao': request.form.get('situacao'),
                'nome_completo': request.form.get('nome_completo'),
                'cpf_cnpj': request.form.get('cpf_cnpj'),
                'telefone': request.form.get('telefone'),
                'email': request.form.get('email'),
                'endereco_completo': request.form.get('endereco_completo'),
                'cidade': request.form.get('cidade'),
                'patrimonio': request.form.get('patrimonio') or None,
                'estado_civil': request.form.get('estado_civil'),
                'nome_conjuge': request.form.get('nome_conjuge'),
                'cpf_conjuge': request.form.get('cpf_conjuge'),
                'observacoes': request.form.get('observacoes'),
            }
            
            if not dados['nome_completo']:
                flash('Nome completo é obrigatório!', 'danger')
                return render_template('pessoas/form.html', pessoa=dados, config=app.config)
            
            pessoa_id = db.insert('pessoas', dados)
            
            if pessoa_id:
                flash(f'Pessoa cadastrada com sucesso! ID: {pessoa_id}', 'success')
                return redirect(url_for('listar_pessoas'))
            else:
                flash('Erro ao cadastrar pessoa.', 'danger')
                
        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')
    
    return render_template('pessoas/form.html', pessoa=None, config=app.config)


@app.route('/pessoas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_pessoa(id):
    """Edita uma pessoa existente."""
    pessoa = db.get_by_id('pessoas', id)

    if not pessoa:
        flash('Pessoa não encontrada.', 'danger')
        return redirect(url_for('listar_pessoas'))

    if request.method == 'POST':
        try:
            dados = {
                'situacao': request.form.get('situacao'),
                'nome_completo': request.form.get('nome_completo'),
                'cpf_cnpj': request.form.get('cpf_cnpj'),
                'telefone': request.form.get('telefone'),
                'email': request.form.get('email'),
                'endereco_completo': request.form.get('endereco_completo'),
                'cidade': request.form.get('cidade'),
                'patrimonio': request.form.get('patrimonio') or None,
                'estado_civil': request.form.get('estado_civil'),
                'nome_conjuge': request.form.get('nome_conjuge'),
                'cpf_conjuge': request.form.get('cpf_conjuge'),
                'observacoes': request.form.get('observacoes'),
            }

            if not dados['nome_completo']:
                flash('Nome completo é obrigatório!', 'danger')
                return render_template('pessoas/form.html', pessoa=dados, config=app.config)

            if db.update('pessoas', dados, 'id = ?', (id,)):
                flash('Pessoa atualizada com sucesso!', 'success')
                return redirect(url_for('listar_pessoas'))
            else:
                flash('Erro ao atualizar pessoa.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    return render_template('pessoas/form.html', pessoa=pessoa, config=app.config)


@app.route('/pessoas/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_pessoa(id):
    """Exclui uma pessoa."""
    try:
        if db.delete('pessoas', 'id = ?', (id,)):
            flash('Pessoa excluída com sucesso!', 'success')
        else:
            flash('Erro ao excluir pessoa.', 'danger')
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('listar_pessoas'))


# ============================================================================
# ROTAS - CONTRATOS
# ============================================================================

@app.route('/contratos')
@login_required
def listar_contratos():
    """Lista todos os contratos."""
    status = request.args.get('status', '')
    
    if status:
        contratos = db.get_where('contratos', 'status_contrato = ?', (status,), 'inicio_contrato DESC')
    else:
        contratos = db.get_all('contratos', 'inicio_contrato DESC')
    
    # Enriquecer com dados relacionados
    for contrato in contratos:
        contrato['imovel'] = db.get_by_id('imoveis', contrato['id_imovel'])
        contrato['inquilino'] = db.get_by_id('pessoas', contrato['id_inquilino'])
        if contrato['id_fiador']:
            contrato['fiador'] = db.get_by_id('pessoas', contrato['id_fiador'])
    
    return render_template('contratos/listar.html', contratos=contratos, status=status, config=app.config)


@app.route('/contratos/novo', methods=['GET', 'POST'])
@login_required
def novo_contrato():
    """Cadastra novo contrato."""
    if request.method == 'POST':
        try:
            # Coletar dados do formulário
            dados = {
                'id_imovel': request.form.get('id_imovel'),
                'id_inquilino': request.form.get('id_inquilino'),
                'id_fiador': request.form.get('id_fiador') or None,
                'garantia': request.form.get('garantia'),
                'inicio_contrato': request.form.get('inicio_contrato'),
                'fim_contrato': request.form.get('fim_contrato') or None,
                'valor_aluguel': request.form.get('valor_aluguel'),
                'dia_vencimento': request.form.get('dia_vencimento'),
                'indice_reajuste': request.form.get('indice_reajuste', 'IGPM'),
                'data_base_reajuste': request.form.get('data_base_reajuste') or None,
                'status_contrato': request.form.get('status_contrato', 'Ativo'),
                'observacoes': request.form.get('observacoes'),
            }
            
            # Validações
            if not dados['id_imovel'] or not dados['id_inquilino']:
                flash('Imóvel e Inquilino são obrigatórios!', 'danger')
                return redirect(url_for('novo_contrato'))
            
            # Se garantia é fiança, fiador é obrigatório
            if dados['garantia'] == 'fiança' and not dados['id_fiador']:
                flash('Fiador é obrigatório quando a garantia é "fiança"!', 'danger')
                return redirect(url_for('novo_contrato'))
            
            # Inserir contrato
            contrato_id = db.insert('contratos', dados)
            
            if contrato_id:
                flash(f'Contrato cadastrado com sucesso! ID: {contrato_id}', 'success')
                return redirect(url_for('listar_contratos'))
            else:
                flash('Erro ao cadastrar contrato.', 'danger')
                
        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')
    
    # Buscar dados para os selects
    imoveis_disponiveis = db.get_imoveis_disponiveis()
    inquilinos = db.get_where('pessoas', "situacao IN ('Inquilino', 'Ambos')", (), 'nome_completo')
    fiadores = db.get_where('pessoas', "situacao IN ('Fiador', 'Ambos')", (), 'nome_completo')
    
    return render_template('contratos/form.html', 
                         contrato=None, 
                         imoveis_disponiveis=imoveis_disponiveis,
                         inquilinos=inquilinos,
                         fiadores=fiadores,
                         config=app.config)


@app.route('/contratos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_contrato(id):
    """Edita um contrato existente."""
    contrato = db.get_by_id('contratos', id)
    
    if not contrato:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('listar_contratos'))
    
    if request.method == 'POST':
        try:
            dados = {
                'id_imovel': request.form.get('id_imovel'),
                'id_inquilino': request.form.get('id_inquilino'),
                'id_fiador': request.form.get('id_fiador') or None,
                'garantia': request.form.get('garantia'),
                'inicio_contrato': request.form.get('inicio_contrato'),
                'fim_contrato': request.form.get('fim_contrato') or None,
                'valor_aluguel': request.form.get('valor_aluguel'),
                'dia_vencimento': request.form.get('dia_vencimento'),
                'indice_reajuste': request.form.get('indice_reajuste'),
                'data_base_reajuste': request.form.get('data_base_reajuste') or None,
                'status_contrato': request.form.get('status_contrato'),
                'observacoes': request.form.get('observacoes'),
            }
            
            # Validação de fiador
            if dados['garantia'] == 'fiança' and not dados['id_fiador']:
                flash('Fiador é obrigatório quando a garantia é "fiança"!', 'danger')
                return redirect(url_for('editar_contrato', id=id))
            
            if db.update('contratos', dados, 'id = ?', (id,)):
                flash('Contrato atualizado com sucesso!', 'success')
                return redirect(url_for('listar_contratos'))
            else:
                flash('Erro ao atualizar contrato.', 'danger')
                
        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')
    
    # Buscar dados para os selects (incluindo imóvel atual)
    imoveis_disponiveis = db.get_imoveis_disponiveis()
    # Adicionar o imóvel atual se não estiver na lista
    imovel_atual = db.get_by_id('imoveis', contrato['id_imovel'])
    if imovel_atual and imovel_atual not in imoveis_disponiveis:
        imoveis_disponiveis.insert(0, imovel_atual)
    
    inquilinos = db.get_where('pessoas', "situacao IN ('Inquilino', 'Ambos')", (), 'nome_completo')
    fiadores = db.get_where('pessoas', "situacao IN ('Fiador', 'Ambos')", (), 'nome_completo')
    
    return render_template('contratos/form.html',
                         contrato=contrato,
                         imoveis_disponiveis=imoveis_disponiveis,
                         inquilinos=inquilinos,
                         fiadores=fiadores,
                         config=app.config)


@app.route('/contratos/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_contrato(id):
    """Exclui um contrato."""
    try:
        if db.delete('contratos', 'id = ?', (id,)):
            flash('Contrato excluído com sucesso!', 'success')
        else:
            flash('Erro ao excluir contrato.', 'danger')
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('listar_contratos'))


# ============================================================================
# ROTAS - DESPESAS
# ============================================================================

@app.route('/despesas')
@login_required
def listar_despesas():
    """Lista todas as despesas."""
    tipo = request.args.get('tipo', '')
    status = request.args.get('status', '')
    
    if tipo:
        despesas = db.get_where('despesas', 'tipo_despesa = ?', (tipo,), 'vencimento_previsto DESC')
    elif status == 'pendente':
        despesas = db.get_despesas_pendentes()
    else:
        despesas = db.execute_query("""
            SELECT d.*, i.endereco_completo as imovel_endereco
            FROM despesas d
            JOIN imoveis i ON d.id_imovel = i.id
            ORDER BY d.vencimento_previsto DESC
            LIMIT 100
        """)
    
    hoje = date.today().strftime('%Y-%m-%d')
    return render_template('despesas/listar.html', despesas=despesas, tipo=tipo, status=status, hoje=hoje, config=app.config)


@app.route('/despesas/nova', methods=['GET', 'POST'])
@login_required
def nova_despesa():
    """Cadastra nova despesa."""
    if request.method == 'POST':
        try:
            dados = {
                'id_imovel': request.form.get('id_imovel'),
                'tipo_despesa': request.form.get('tipo_despesa'),
                'motivo_despesa': request.form.get('motivo_despesa'),
                'mes_referencia': request.form.get('mes_referencia') or None,
                'valor_previsto': request.form.get('valor_previsto'),
                'valor_pago': request.form.get('valor_pago') or None,
                'vencimento_previsto': request.form.get('vencimento_previsto') or None,
                'data_pagamento': request.form.get('data_pagamento') or None,
                'observacoes': request.form.get('observacoes'),
            }

            # Validações
            if not dados['id_imovel'] or not dados['tipo_despesa'] or not dados['valor_previsto']:
                flash('Imóvel, Tipo de Despesa e Valor Previsto são obrigatórios!', 'danger')
                return redirect(url_for('nova_despesa'))

            despesa_id = db.insert('despesas', dados)

            if despesa_id:
                flash(f'Despesa cadastrada com sucesso! ID: {despesa_id}', 'success')
                return redirect(url_for('listar_despesas'))
            else:
                flash('Erro ao cadastrar despesa.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    # Buscar imóveis para o select
    imoveis = db.get_all('imoveis', 'endereco_completo')

    return render_template('despesas/form.html', despesa=None, imoveis=imoveis, config=app.config)


@app.route('/despesas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_despesa(id):
    """Edita uma despesa existente."""
    despesa = db.get_by_id('despesas', id)

    if not despesa:
        flash('Despesa não encontrada.', 'danger')
        return redirect(url_for('listar_despesas'))

    if request.method == 'POST':
        try:
            dados = {
                'id_imovel': request.form.get('id_imovel'),
                'tipo_despesa': request.form.get('tipo_despesa'),
                'motivo_despesa': request.form.get('motivo_despesa'),
                'mes_referencia': request.form.get('mes_referencia') or None,
                'valor_previsto': request.form.get('valor_previsto'),
                'valor_pago': request.form.get('valor_pago') or None,
                'vencimento_previsto': request.form.get('vencimento_previsto') or None,
                'data_pagamento': request.form.get('data_pagamento') or None,
                'observacoes': request.form.get('observacoes'),
            }

            if db.update('despesas', dados, 'id = ?', (id,)):
                flash('Despesa atualizada com sucesso!', 'success')
                return redirect(url_for('listar_despesas'))
            else:
                flash('Erro ao atualizar despesa.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    # Buscar imóveis para o select
    imoveis = db.get_all('imoveis', 'endereco_completo')

    return render_template('despesas/form.html', despesa=despesa, imoveis=imoveis, config=app.config)


@app.route('/despesas/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_despesa(id):
    """Exclui uma despesa."""
    try:
        if db.delete('despesas', 'id = ?', (id,)):
            flash('Despesa excluída com sucesso!', 'success')
        else:
            flash('Erro ao excluir despesa.', 'danger')
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('listar_despesas'))


@app.route('/despesas/<int:id>/pagar', methods=['POST'])
@login_required
def pagar_despesa(id):
    """Marca uma despesa como paga."""
    try:
        despesa = db.get_by_id('despesas', id)
        if not despesa:
            flash('Despesa não encontrada.', 'danger')
            return redirect(url_for('listar_despesas'))

        dados = {
            'valor_pago': despesa['valor_previsto'],
            'data_pagamento': date.today().strftime('%Y-%m-%d')
        }

        if db.update('despesas', dados, 'id = ?', (id,)):
            flash('Despesa marcada como paga!', 'success')
        else:
            flash('Erro ao atualizar despesa.', 'danger')

    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('listar_despesas'))


@app.route('/despesas/gerar-iptu-anual', methods=['POST'])
@login_required
def gerar_iptu_anual():
    """Gera despesas de IPTU anual para todos os imóveis."""
    try:
        # Receber data de vencimento do formulário
        data_vencimento = request.form.get('data_vencimento')
        if not data_vencimento:
            flash('Data de vencimento é obrigatória.', 'danger')
            return redirect(url_for('listar_despesas'))

        ano_atual = data_vencimento[:4]  # Extrair ano da data informada
        vencimento = data_vencimento
        mes_referencia = f"{ano_atual}-01-01"  # Janeiro do ano do vencimento

        # Buscar imóveis com IPTU cadastrado
        imoveis = db.execute_query("""
            SELECT id, endereco_completo, valor_iptu_anual
            FROM imoveis
            WHERE valor_iptu_anual IS NOT NULL AND valor_iptu_anual > 0
        """)

        if not imoveis:
            flash('Nenhum imóvel com valor de IPTU cadastrado.', 'warning')
            return redirect(url_for('listar_despesas'))

        contador = 0
        ignorados = 0

        for imovel in imoveis:
            # Verificar se já existe IPTU para este imóvel neste ano
            existente = db.execute_query("""
                SELECT id FROM despesas
                WHERE id_imovel = ? AND tipo_despesa = 'IPTU'
                AND strftime('%Y', mes_referencia) = ?
            """, (imovel['id'], str(ano_atual)))

            if existente:
                ignorados += 1
                continue

            # Criar despesa de IPTU
            dados = {
                'id_imovel': imovel['id'],
                'tipo_despesa': 'IPTU',
                'motivo_despesa': f'IPTU Anual {ano_atual}',
                'mes_referencia': mes_referencia,
                'valor_previsto': imovel['valor_iptu_anual'],
                'vencimento_previsto': vencimento,
            }

            if db.insert('despesas', dados):
                contador += 1

        if contador > 0:
            # Formatar data para exibição (DD/MM/AAAA)
            data_formatada = f"{vencimento[8:10]}/{vencimento[5:7]}/{vencimento[:4]}"
            flash(f'IPTU gerado com sucesso para {contador} imóvel(is)! Vencimento: {data_formatada}', 'success')
        if ignorados > 0:
            flash(f'{ignorados} imóvel(is) já tinham IPTU lançado para {ano_atual}.', 'info')

    except Exception as e:
        flash(f'Erro ao gerar IPTU: {str(e)}', 'danger')

    return redirect(url_for('listar_despesas'))


@app.route('/despesas/gerar-condominio-mensal', methods=['POST'])
@login_required
def gerar_condominio_mensal():
    """Gera despesas de condomínio mensal para todos os imóveis com condomínio > 0."""
    try:
        hoje = date.today()
        ano_atual = hoje.year
        mes_atual = hoje.month
        mes_referencia = f"{ano_atual}-{mes_atual:02d}-01"

        # Buscar imóveis com condomínio cadastrado
        imoveis = db.execute_query("""
            SELECT id, endereco_completo, condominio_sugerido, dia_venc_condominio
            FROM imoveis
            WHERE condominio_sugerido IS NOT NULL AND condominio_sugerido > 0
        """)

        if not imoveis:
            flash('Nenhum imóvel com valor de condomínio cadastrado.', 'warning')
            return redirect(url_for('listar_despesas'))

        contador = 0
        ignorados = 0

        for imovel in imoveis:
            # Verificar se já existe condomínio para este imóvel neste mês
            existente = db.execute_query("""
                SELECT id FROM despesas
                WHERE id_imovel = ? AND tipo_despesa = 'Condomínio'
                AND strftime('%Y-%m', mes_referencia) = ?
            """, (imovel['id'], f"{ano_atual}-{mes_atual:02d}"))

            if existente:
                ignorados += 1
                continue

            # Calcular data de vencimento
            dia_venc = imovel['dia_venc_condominio'] or 10
            # Ajustar dia se for maior que o último dia do mês
            import calendar
            ultimo_dia = calendar.monthrange(ano_atual, mes_atual)[1]
            dia_venc = min(dia_venc, ultimo_dia)
            vencimento = f"{ano_atual}-{mes_atual:02d}-{dia_venc:02d}"

            # Criar despesa de condomínio
            dados = {
                'id_imovel': imovel['id'],
                'tipo_despesa': 'Condomínio',
                'motivo_despesa': f'Condomínio {mes_atual:02d}/{ano_atual}',
                'mes_referencia': mes_referencia,
                'valor_previsto': imovel['condominio_sugerido'],
                'vencimento_previsto': vencimento,
            }

            if db.insert('despesas', dados):
                contador += 1

        mes_nome = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes_atual]

        if contador > 0:
            flash(f'Condomínio gerado com sucesso para {contador} imóvel(is)! Referência: {mes_nome}/{ano_atual}', 'success')
        if ignorados > 0:
            flash(f'{ignorados} imóvel(is) já tinham condomínio lançado para {mes_nome}/{ano_atual}.', 'info')

    except Exception as e:
        flash(f'Erro ao gerar condomínio: {str(e)}', 'danger')

    return redirect(url_for('listar_despesas'))


# ============================================================================
# ROTAS - RECEITAS
# ============================================================================

@app.route('/receitas')
@login_required
def listar_receitas():
    """Lista todas as receitas."""
    status = request.args.get('status', '')
    
    if status == 'pendente':
        receitas = db.get_receitas_pendentes()
    else:
        receitas = db.execute_query("""
            SELECT r.*, c.valor_aluguel, i.endereco_completo as imovel_endereco,
                   p.nome_completo as inquilino_nome
            FROM receitas r
            JOIN contratos c ON r.id_contrato = c.id
            JOIN imoveis i ON c.id_imovel = i.id
            JOIN pessoas p ON c.id_inquilino = p.id
            ORDER BY r.vencimento_previsto DESC
            LIMIT 100
        """)
    
    return render_template('receitas/listar.html', receitas=receitas, status=status, config=app.config)


@app.route('/receitas/gerar-faturamento-mensal', methods=['POST'])
@login_required
def gerar_faturamento_mensal():
    """Gera receitas de aluguel para todos os contratos ativos no mês corrente."""
    try:
        import calendar
        hoje = date.today()
        ano_atual = hoje.year
        mes_atual = hoje.month
        mes_referencia = f"{ano_atual}-{mes_atual:02d}-01"

        # Buscar contratos ativos ou prorrogados
        contratos = db.execute_query("""
            SELECT c.id, c.valor_aluguel, c.dia_vencimento,
                   i.endereco_completo as imovel_endereco,
                   i.condominio_sugerido, i.valor_iptu_anual, i.forma_pagamento_iptu,
                   p.nome_completo as inquilino_nome
            FROM contratos c
            JOIN imoveis i ON c.id_imovel = i.id
            JOIN pessoas p ON c.id_inquilino = p.id
            WHERE c.status_contrato IN ('Ativo', 'Prorrogado')
        """)

        if not contratos:
            flash('Nenhum contrato ativo encontrado.', 'warning')
            return redirect(url_for('listar_receitas'))

        contador = 0
        ignorados = 0

        for contrato in contratos:
            # Verificar se já existe receita para este contrato neste mês
            existente = db.execute_query("""
                SELECT id FROM receitas
                WHERE id_contrato = ?
                AND strftime('%Y-%m', mes_referencia) = ?
            """, (contrato['id'], f"{ano_atual}-{mes_atual:02d}"))

            if existente:
                ignorados += 1
                continue

            # Calcular data de vencimento
            dia_venc = contrato['dia_vencimento'] or 10
            ultimo_dia = calendar.monthrange(ano_atual, mes_atual)[1]
            dia_venc = min(dia_venc, ultimo_dia)
            vencimento = f"{ano_atual}-{mes_atual:02d}-{dia_venc:02d}"

            # Valores
            aluguel = contrato['valor_aluguel'] or 0
            condominio = contrato['condominio_sugerido'] or 0

            # IPTU mensal (se forma de pagamento for mensal)
            iptu = 0
            if contrato['forma_pagamento_iptu'] == 'Mensal' and contrato['valor_iptu_anual']:
                iptu = round(contrato['valor_iptu_anual'] / 12, 2)

            valor_total = aluguel + condominio + iptu

            # Criar receita
            dados = {
                'id_contrato': contrato['id'],
                'mes_referencia': mes_referencia,
                'aluguel_devido': aluguel,
                'condominio_devido': condominio if condominio > 0 else None,
                'iptu_devido': iptu if iptu > 0 else None,
                'desconto_multa': None,
                'valor_total_devido': valor_total,
                'vencimento_previsto': vencimento,
                'status': 'Pendente',
                'observacoes': f'Faturamento automático {mes_atual:02d}/{ano_atual}',
            }

            if db.insert('receitas', dados):
                contador += 1

        mes_nome = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes_atual]

        if contador > 0:
            flash(f'Faturamento gerado com sucesso para {contador} contrato(s)! Referência: {mes_nome}/{ano_atual}', 'success')
        if ignorados > 0:
            flash(f'{ignorados} contrato(s) já tinham faturamento lançado para {mes_nome}/{ano_atual}.', 'info')

    except Exception as e:
        flash(f'Erro ao gerar faturamento: {str(e)}', 'danger')

    return redirect(url_for('listar_receitas'))


@app.route('/receitas/nova', methods=['GET', 'POST'])
@login_required
def nova_receita():
    """Cadastra nova receita."""
    if request.method == 'POST':
        try:
            # Calcular valor total
            aluguel = float(request.form.get('aluguel_devido') or 0)
            condominio = float(request.form.get('condominio_devido') or 0)
            iptu = float(request.form.get('iptu_devido') or 0)
            desconto_multa = float(request.form.get('desconto_multa') or 0)
            valor_total = aluguel + condominio + iptu + desconto_multa

            dados = {
                'id_contrato': request.form.get('id_contrato'),
                'mes_referencia': request.form.get('mes_referencia'),
                'aluguel_devido': aluguel,
                'condominio_devido': condominio or None,
                'iptu_devido': iptu or None,
                'desconto_multa': desconto_multa or None,
                'valor_total_devido': valor_total,
                'vencimento_previsto': request.form.get('vencimento_previsto'),
                'data_recebimento': request.form.get('data_recebimento') or None,
                'valor_recebido': request.form.get('valor_recebido') or None,
                'status': request.form.get('status', 'Pendente'),
                'observacoes': request.form.get('observacoes'),
            }

            # Validações
            if not dados['id_contrato'] or not dados['mes_referencia'] or not dados['vencimento_previsto']:
                flash('Contrato, Mês de Referência e Vencimento são obrigatórios!', 'danger')
                return redirect(url_for('nova_receita'))

            receita_id = db.insert('receitas', dados)

            if receita_id:
                flash(f'Receita cadastrada com sucesso! ID: {receita_id}', 'success')
                return redirect(url_for('listar_receitas'))
            else:
                flash('Erro ao cadastrar receita.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    # Buscar contratos ativos para o select
    contratos = db.execute_query("""
        SELECT c.id, c.valor_aluguel, c.dia_vencimento,
               i.endereco_completo as imovel_endereco,
               p.nome_completo as inquilino_nome
        FROM contratos c
        JOIN imoveis i ON c.id_imovel = i.id
        JOIN pessoas p ON c.id_inquilino = p.id
        WHERE c.status_contrato IN ('Ativo', 'Prorrogado')
        ORDER BY i.endereco_completo
    """)

    return render_template('receitas/form.html', receita=None, contratos=contratos, config=app.config)


@app.route('/receitas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_receita(id):
    """Edita uma receita existente."""
    receita = db.get_by_id('receitas', id)

    if not receita:
        flash('Receita não encontrada.', 'danger')
        return redirect(url_for('listar_receitas'))

    if request.method == 'POST':
        try:
            # Calcular valor total
            aluguel = float(request.form.get('aluguel_devido') or 0)
            condominio = float(request.form.get('condominio_devido') or 0)
            iptu = float(request.form.get('iptu_devido') or 0)
            desconto_multa = float(request.form.get('desconto_multa') or 0)
            valor_total = aluguel + condominio + iptu + desconto_multa

            dados = {
                'id_contrato': request.form.get('id_contrato'),
                'mes_referencia': request.form.get('mes_referencia'),
                'aluguel_devido': aluguel,
                'condominio_devido': condominio or None,
                'iptu_devido': iptu or None,
                'desconto_multa': desconto_multa or None,
                'valor_total_devido': valor_total,
                'vencimento_previsto': request.form.get('vencimento_previsto'),
                'data_recebimento': request.form.get('data_recebimento') or None,
                'valor_recebido': request.form.get('valor_recebido') or None,
                'status': request.form.get('status'),
                'observacoes': request.form.get('observacoes'),
            }

            if db.update('receitas', dados, 'id = ?', (id,)):
                flash('Receita atualizada com sucesso!', 'success')
                return redirect(url_for('listar_receitas'))
            else:
                flash('Erro ao atualizar receita.', 'danger')

        except Exception as e:
            flash(f'Erro: {str(e)}', 'danger')

    # Buscar contratos ativos para o select
    contratos = db.execute_query("""
        SELECT c.id, c.valor_aluguel, c.dia_vencimento,
               i.endereco_completo as imovel_endereco,
               p.nome_completo as inquilino_nome
        FROM contratos c
        JOIN imoveis i ON c.id_imovel = i.id
        JOIN pessoas p ON c.id_inquilino = p.id
        WHERE c.status_contrato IN ('Ativo', 'Prorrogado')
           OR c.id = ?
        ORDER BY i.endereco_completo
    """, (receita['id_contrato'],))

    return render_template('receitas/form.html', receita=receita, contratos=contratos, config=app.config)


@app.route('/receitas/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_receita(id):
    """Exclui uma receita."""
    try:
        if db.delete('receitas', 'id = ?', (id,)):
            flash('Receita excluída com sucesso!', 'success')
        else:
            flash('Erro ao excluir receita.', 'danger')
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('listar_receitas'))


@app.route('/receitas/<int:id>/receber', methods=['POST'])
@login_required
def receber_receita(id):
    """Marca uma receita como recebida."""
    try:
        receita = db.get_by_id('receitas', id)
        if not receita:
            flash('Receita não encontrada.', 'danger')
            return redirect(url_for('listar_receitas'))

        dados = {
            'valor_recebido': receita['valor_total_devido'],
            'data_recebimento': date.today().strftime('%Y-%m-%d'),
            'status': 'Recebido'
        }

        if db.update('receitas', dados, 'id = ?', (id,)):
            flash('Receita marcada como recebida!', 'success')
        else:
            flash('Erro ao atualizar receita.', 'danger')

    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('listar_receitas'))


# ============================================================================
# ROTAS - RELATÓRIOS
# ============================================================================

@app.route('/relatorios')
@login_required
def listar_relatorios():
    """Página de relatórios."""
    return render_template('relatorios/index.html')


@app.route('/relatorios/despesas-pendentes')
@login_required
def relatorio_despesas_pendentes():
    """Relatório de despesas pendentes."""
    # Filtros
    filtro_tipo = request.args.get('tipo', '')
    filtro_vencimento = request.args.get('vencimento_ate', '')
    filtro_situacao = request.args.get('situacao', 'todas')

    hoje = date.today()

    # Query base - despesas não pagas
    query = """
        SELECT d.*, i.endereco_completo as imovel_endereco
        FROM despesas d
        LEFT JOIN imoveis i ON d.id_imovel = i.id
        WHERE d.data_pagamento IS NULL
    """
    params = []

    # Aplicar filtro de tipo
    if filtro_tipo:
        query += " AND d.tipo_despesa = ?"
        params.append(filtro_tipo)

    # Aplicar filtro de vencimento
    if filtro_vencimento:
        query += " AND d.vencimento_previsto <= ?"
        params.append(filtro_vencimento)

    # Aplicar filtro de situação
    if filtro_situacao == 'vencidas':
        query += " AND d.vencimento_previsto < ?"
        params.append(hoje.strftime('%Y-%m-%d'))
    elif filtro_situacao == 'a_vencer':
        query += " AND d.vencimento_previsto >= ?"
        params.append(hoje.strftime('%Y-%m-%d'))

    query += " ORDER BY d.vencimento_previsto ASC"

    despesas = db.execute_query(query, tuple(params))

    # Adicionar flag de vencida e calcular estatísticas
    total_valor = 0
    total_vencidas = 0
    total_a_vencer = 0

    for despesa in despesas:
        venc = despesa.get('vencimento_previsto')
        if venc:
            try:
                venc_date = datetime.strptime(venc, '%Y-%m-%d').date()
                despesa['vencida'] = venc_date < hoje
                if despesa['vencida']:
                    total_vencidas += 1
                else:
                    total_a_vencer += 1
            except:
                despesa['vencida'] = False
                total_a_vencer += 1
        else:
            despesa['vencida'] = False
            total_a_vencer += 1

        total_valor += despesa.get('valor_previsto', 0) or 0

    return render_template('relatorios/despesas_pendentes.html',
                         despesas=despesas,
                         total_despesas=len(despesas),
                         total_valor=total_valor,
                         total_vencidas=total_vencidas,
                         total_a_vencer=total_a_vencer,
                         filtro_tipo=filtro_tipo,
                         filtro_vencimento=filtro_vencimento,
                         filtro_situacao=filtro_situacao)


@app.route('/relatorios/despesas-pendentes/excel')
@login_required
def relatorio_despesas_pendentes_excel():
    """Exporta relatório de despesas pendentes para Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Filtros
    filtro_tipo = request.args.get('tipo', '')
    filtro_vencimento = request.args.get('vencimento_ate', '')
    filtro_situacao = request.args.get('situacao', 'todas')

    hoje = date.today()

    # Query base - despesas não pagas
    query = """
        SELECT d.*, i.endereco_completo as imovel_endereco
        FROM despesas d
        LEFT JOIN imoveis i ON d.id_imovel = i.id
        WHERE d.data_pagamento IS NULL
    """
    params = []

    if filtro_tipo:
        query += " AND d.tipo_despesa = ?"
        params.append(filtro_tipo)

    if filtro_vencimento:
        query += " AND d.vencimento_previsto <= ?"
        params.append(filtro_vencimento)

    if filtro_situacao == 'vencidas':
        query += " AND d.vencimento_previsto < ?"
        params.append(hoje.strftime('%Y-%m-%d'))
    elif filtro_situacao == 'a_vencer':
        query += " AND d.vencimento_previsto >= ?"
        params.append(hoje.strftime('%Y-%m-%d'))

    query += " ORDER BY d.vencimento_previsto ASC"

    despesas = db.execute_query(query, tuple(params))

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Despesas Pendentes"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f"RELATÓRIO DE DESPESAS PENDENTES - Gerado em {hoje.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Cabeçalhos
    headers = ['ID', 'Imóvel', 'Tipo', 'Descrição', 'Vencimento', 'Valor', 'Situação']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Dados
    total_valor = 0
    for row_num, despesa in enumerate(despesas, 4):
        venc = despesa.get('vencimento_previsto')
        try:
            venc_date = datetime.strptime(venc, '%Y-%m-%d').date() if venc else None
            situacao = "VENCIDA" if venc_date and venc_date < hoje else "A Vencer"
        except:
            situacao = "A Vencer"

        venc_formatado = datetime.strptime(venc, '%Y-%m-%d').strftime('%d/%m/%Y') if venc else ""
        valor = despesa.get('valor_previsto', 0) or 0
        total_valor += valor

        dados = [
            despesa.get('id'),
            despesa.get('imovel_endereco', ''),
            despesa.get('tipo_despesa', ''),
            despesa.get('motivo_despesa', ''),
            venc_formatado,
            valor,
            situacao
        ]

        for col, valor_celula in enumerate(dados, 1):
            cell = ws.cell(row=row_num, column=col, value=valor_celula)
            cell.border = border
            if col == 6:  # Valor
                cell.number_format = 'R$ #,##0.00'

    # Total
    row_total = len(despesas) + 4
    ws.cell(row=row_total, column=5, value="TOTAL:").font = Font(bold=True)
    total_cell = ws.cell(row=row_total, column=6, value=total_valor)
    total_cell.font = Font(bold=True)
    total_cell.number_format = 'R$ #,##0.00'

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nome_arquivo = f'despesas_pendentes_{timestamp}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )


# ============================================================================
# ROTAS - EXPORTAR E IMPORTAR DADOS
# ============================================================================

@app.route('/dados')
@login_required
def pagina_dados():
    """Pagina de exportacao e importacao de dados."""
    ultimo_backup = obter_ultimo_backup()
    return render_template('dados/index.html', ultimo_backup=ultimo_backup)


@app.route('/dados/exportar')
@login_required
def exportar_dados():
    """Exporta todas as tabelas para um arquivo ZIP com CSVs."""
    try:
        # Criar arquivo ZIP em memoria
        zip_buffer = io.BytesIO()

        tabelas = ['imoveis', 'pessoas', 'contratos', 'despesas', 'receitas']

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for tabela in tabelas:
                # Buscar dados da tabela
                dados = db.get_all(tabela)

                if not dados:
                    continue

                # Criar CSV em memoria
                csv_buffer = io.StringIO()
                cabecalhos = list(dados[0].keys())

                writer = csv.DictWriter(csv_buffer, fieldnames=cabecalhos)
                writer.writeheader()
                writer.writerows(dados)

                # Adicionar ao ZIP
                csv_content = csv_buffer.getvalue().encode('utf-8')
                zip_file.writestr(f'{tabela}.csv', csv_content)

            # Adicionar arquivo de instrucoes
            instrucoes = """IMOBIPRO - GUIA DE IMPORTACAO
==============================

Este arquivo ZIP contem os dados exportados do sistema ImobiPro.

ARQUIVOS INCLUIDOS:
- imoveis.csv      : Cadastro de imoveis
- pessoas.csv      : Cadastro de pessoas (inquilinos/fiadores)
- contratos.csv    : Contratos de locacao
- despesas.csv     : Despesas dos imoveis
- receitas.csv     : Receitas de alugueis

FORMATO DOS ARQUIVOS:
- Todos os arquivos estao em formato CSV (valores separados por virgula)
- Codificacao: UTF-8
- Primeira linha contem os cabecalhos (nomes das colunas)
- Datas no formato: AAAA-MM-DD (ex: 2026-01-20)
- Valores decimais com ponto (ex: 1500.50)

PARA REIMPORTAR:
1. Acesse a pagina Dados > Importar
2. Selecione este arquivo ZIP
3. Confirme a importacao

ATENCAO:
- A importacao SUBSTITUI todos os dados existentes
- Faca backup antes de importar
- Mantenha a estrutura dos arquivos CSV
- Nao altere os nomes das colunas

ORDEM DE IMPORTACAO (automatica):
1. imoveis (base para contratos e despesas)
2. pessoas (base para contratos)
3. contratos (depende de imoveis e pessoas)
4. despesas (depende de imoveis)
5. receitas (depende de contratos)

Em caso de duvidas, consulte a documentacao do sistema.
"""
            zip_file.writestr('LEIA-ME.txt', instrucoes.encode('utf-8'))

        # Preparar para download
        zip_buffer.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f'imobipro_backup_{timestamp}.zip'

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        flash(f'Erro ao exportar dados: {str(e)}', 'danger')
        return redirect(url_for('pagina_dados'))


@app.route('/dados/importar', methods=['POST'])
@login_required
def importar_dados():
    """Importa dados de um arquivo ZIP com CSVs."""
    try:
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo enviado.', 'danger')
            return redirect(url_for('pagina_dados'))

        arquivo = request.files['arquivo']

        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'danger')
            return redirect(url_for('pagina_dados'))

        if not arquivo.filename.endswith('.zip'):
            flash('O arquivo deve ser um ZIP.', 'danger')
            return redirect(url_for('pagina_dados'))

        # Ordem de importacao (respeitar foreign keys)
        ordem_tabelas = ['imoveis', 'pessoas', 'contratos', 'despesas', 'receitas']
        resultados = []

        with zipfile.ZipFile(arquivo, 'r') as zip_file:
            arquivos_no_zip = zip_file.namelist()

            for tabela in ordem_tabelas:
                nome_csv = f'{tabela}.csv'

                if nome_csv not in arquivos_no_zip:
                    resultados.append(f'{tabela}: arquivo nao encontrado')
                    continue

                # Ler CSV do ZIP
                with zip_file.open(nome_csv) as csv_file:
                    # Decodificar bytes para string
                    conteudo = csv_file.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(conteudo))
                    dados = list(reader)

                if not dados:
                    resultados.append(f'{tabela}: vazio')
                    continue

                # Obter colunas da tabela
                conn = db.connect()
                cursor = conn.cursor()
                cursor.execute(f"PRAGMA table_info({tabela})")
                colunas_tabela = [row[1] for row in cursor.fetchall()]

                # Desabilitar foreign keys temporariamente
                cursor.execute("PRAGMA foreign_keys = OFF")

                # Limpar tabela
                cursor.execute(f"DELETE FROM {tabela}")

                # Filtrar colunas validas
                colunas_csv = list(dados[0].keys())
                colunas_validas = [c for c in colunas_csv if c in colunas_tabela]

                # Preparar INSERT
                placeholders = ', '.join(['?' for _ in colunas_validas])
                colunas_str = ', '.join(colunas_validas)
                query = f"INSERT INTO {tabela} ({colunas_str}) VALUES ({placeholders})"

                # Inserir dados
                contador = 0
                for registro in dados:
                    valores = []
                    for coluna in colunas_validas:
                        valor = registro.get(coluna, '')
                        if valor == '' or valor == 'None':
                            valor = None
                        valores.append(valor)

                    try:
                        cursor.execute(query, valores)
                        contador += 1
                    except Exception as e:
                        print(f"Erro ao inserir em {tabela}: {e}")

                # Reabilitar foreign keys
                cursor.execute("PRAGMA foreign_keys = ON")
                conn.commit()
                conn.close()

                resultados.append(f'{tabela}: {contador} registros')

        flash(f'Importacao concluida! ' + ', '.join(resultados), 'success')

    except zipfile.BadZipFile:
        flash('Arquivo ZIP invalido ou corrompido.', 'danger')
    except Exception as e:
        flash(f'Erro ao importar: {str(e)}', 'danger')

    return redirect(url_for('pagina_dados'))


@app.route('/dados/executar-backup', methods=['POST'])
@login_required
def executar_backup():
    """Executa backup do banco de dados e salva na pasta backups/."""
    try:
        sucesso, resultado = backup_system.backup_sqlite()
        if sucesso:
            flash(f'Backup criado com sucesso! Arquivo: {os.path.basename(resultado)}', 'success')
        else:
            flash(f'Erro ao criar backup: {resultado}', 'danger')
    except Exception as e:
        flash(f'Erro ao criar backup: {str(e)}', 'danger')

    return redirect(url_for('pagina_dados'))


@app.route('/dados/backup')
@login_required
def backup_banco():
    """Faz download do arquivo de banco de dados SQLite."""
    try:
        import shutil

        # Caminho do banco de dados
        db_path = os.path.join(os.path.dirname(__file__), 'database', 'imobipro.db')

        if not os.path.exists(db_path):
            flash('Arquivo de banco de dados não encontrado.', 'danger')
            return redirect(url_for('pagina_dados'))

        # Criar cópia temporária para evitar problemas de lock
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_dir = tempfile.mkdtemp()
        nome_backup = f'imobipro_backup_{timestamp}.db'
        temp_path = os.path.join(temp_dir, nome_backup)

        # Copiar banco de dados
        shutil.copy2(db_path, temp_path)

        # Enviar arquivo para download
        return send_file(
            temp_path,
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name=nome_backup
        )

    except Exception as e:
        flash(f'Erro ao fazer backup: {str(e)}', 'danger')
        return redirect(url_for('pagina_dados'))


# ============================================================================
# CONTEXTO GLOBAL PARA TEMPLATES
# ============================================================================

@app.context_processor
def inject_globals():
    """Injeta variáveis globais em todos os templates."""
    return {
        'ano_atual': datetime.now().year,
        'app_name': 'ImobiPro',
        'app_version': '1.0.0'
    }


# ============================================================================
# TRATAMENTO DE ERROS
# ============================================================================

@app.errorhandler(404)
def page_not_found(e):
    """Página não encontrada."""
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_server_error(e):
    """Erro interno do servidor."""
    return render_template('errors/500.html'), 500


# ============================================================================
# EXECUÇÃO DA APLICAÇÃO
# ============================================================================

if __name__ == '__main__':
    # Criar diretórios necessários
    os.makedirs('static/uploads', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    
    print("\n" + "="*70)
    print("🏠 IMOBIPRO - SISTEMA DE GESTÃO IMOBILIÁRIA")
    print("="*70)
    print(f"\n✓ Servidor iniciado!")
    print(f"✓ Acesse: http://localhost:5000")
    print(f"✓ Pressione Ctrl+C para encerrar\n")
    
    # Executar aplicação
    app.run(host='0.0.0.0', port=5000, debug=True)