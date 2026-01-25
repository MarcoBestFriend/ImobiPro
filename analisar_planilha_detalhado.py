"""
================================================================================
ANALISADOR DETALHADO DA PLANILHA IMOBIPRO
================================================================================
Descrição: Analisa a planilha em profundidade para identificar:
           - Nomes exatos dos cabeçalhos
           - Tipos de dados
           - Valores únicos
           - Comentários nas células
           - Regras de validação
           - Padrões de dados
================================================================================
"""

from openpyxl import load_workbook
from openpyxl.comments import Comment
import json

def analisar_planilha_completa(caminho_excel):
    """
    Analisa a planilha em detalhes e gera relatório completo.
    """
    print("\n" + "="*80)
    print("ANÁLISE DETALHADA DA PLANILHA IMOBIPRO")
    print("="*80)
    
    # Carregar workbook (com fórmulas e comentários)
    wb = load_workbook(caminho_excel, data_only=False)
    
    print(f"\n📁 Arquivo: {caminho_excel}")
    print(f"📊 Total de abas: {len(wb.sheetnames)}")
    print(f"📋 Abas encontradas: {', '.join(wb.sheetnames)}")
    
    analise_completa = {}
    
    # Analisar cada aba
    for sheet_name in wb.sheetnames:
        print("\n" + "="*80)
        print(f"ABA: {sheet_name}")
        print("="*80)
        
        ws = wb[sheet_name]
        
        # Informações básicas
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"\n📏 Dimensões: {max_row} linhas x {max_col} colunas")
        
        # Analisar cabeçalhos (linha 1)
        print("\n🏷️  CABEÇALHOS (Linha 1):")
        print("-" * 80)
        
        headers = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            header_value = cell.value
            headers.append(header_value)
            
            # Verificar comentários no cabeçalho
            comentario = ""
            if cell.comment:
                comentario = f" [COMENTÁRIO: {cell.comment.text}]"
            
            # Verificar validação de dados
            validacao = ""
            if cell.data_type == 'f':  # Fórmula
                validacao = f" [FÓRMULA: {cell.value}]"
            
            print(f"  Col {col:2d}: '{header_value}'{comentario}{validacao}")
        
        # Analisar tipos de dados e valores únicos
        print("\n📊 ANÁLISE DE DADOS POR COLUNA:")
        print("-" * 80)
        
        analise_colunas = {}
        
        for col_idx, header in enumerate(headers, start=1):
            if not header:
                continue
            
            valores = []
            tipos = set()
            tem_comentarios = False
            tem_formulas = False
            valores_unicos = set()
            
            # Coletar dados da coluna
            for row in range(2, min(max_row + 1, 102)):  # Limitar a 100 linhas
                cell = ws.cell(row=row, column=col_idx)
                valor = cell.value
                
                if valor is not None:
                    valores.append(valor)
                    tipos.add(type(valor).__name__)
                    valores_unicos.add(str(valor))
                
                if cell.comment:
                    tem_comentarios = True
                
                if cell.data_type == 'f':
                    tem_formulas = True
            
            # Relatório da coluna
            print(f"\n  📌 {header}")
            print(f"     Tipo(s) de dado: {', '.join(tipos)}")
            print(f"     Valores não vazios: {len(valores)}")
            print(f"     Valores únicos: {len(valores_unicos)}")
            
            if tem_comentarios:
                print(f"     ⚠️  Contém comentários")
            
            if tem_formulas:
                print(f"     📐 Contém fórmulas")
            
            # Mostrar valores únicos se forem poucos (categorias)
            if len(valores_unicos) <= 10 and len(valores_unicos) > 0:
                print(f"     Valores possíveis: {', '.join(sorted(valores_unicos))}")
            
            # Mostrar amostra de dados
            if len(valores) > 0:
                amostra = valores[:3]
                print(f"     Amostra: {amostra}")
            
            analise_colunas[header] = {
                'tipos': list(tipos),
                'total_valores': len(valores),
                'valores_unicos': len(valores_unicos),
                'tem_comentarios': tem_comentarios,
                'tem_formulas': tem_formulas,
                'categorias': list(valores_unicos) if len(valores_unicos) <= 10 else []
            }
        
        # Procurar comentários em células de dados
        print("\n💬 COMENTÁRIOS ENCONTRADOS:")
        print("-" * 80)
        
        comentarios_encontrados = []
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 50)):
            for cell in row:
                if cell.comment:
                    comentarios_encontrados.append({
                        'celula': cell.coordinate,
                        'valor': cell.value,
                        'comentario': cell.comment.text
                    })
                    print(f"  {cell.coordinate}: {cell.value}")
                    print(f"    → {cell.comment.text}")
        
        if not comentarios_encontrados:
            print("  Nenhum comentário encontrado nas primeiras 50 linhas")
        
        # Procurar validações de dados
        print("\n✓ VALIDAÇÕES DE DADOS:")
        print("-" * 80)
        
        validacoes_encontradas = []
        for col in range(1, max_col + 1):
            for row in range(1, min(max_row + 1, 50)):
                cell = ws.cell(row=row, column=col)
                
                # Verificar se há validação (listas suspensas, etc)
                if hasattr(cell, 'data_validation'):
                    validacoes_encontradas.append({
                        'celula': cell.coordinate,
                        'tipo': 'data_validation'
                    })
        
        if validacoes_encontradas:
            for val in validacoes_encontradas[:10]:  # Mostrar até 10
                print(f"  {val['celula']}: {val['tipo']}")
        else:
            print("  Nenhuma validação explícita encontrada")
        
        # Identificar padrões
        print("\n🔍 PADRÕES IDENTIFICADOS:")
        print("-" * 80)
        
        # Verificar se há IDs sequenciais
        if headers and headers[0] and 'ID' in str(headers[0]).upper():
            ids = []
            for row in range(2, min(max_row + 1, 20)):
                val = ws.cell(row=row, column=1).value
                if val:
                    ids.append(val)
            
            if ids:
                print(f"  ✓ Primeira coluna parece ser ID (valores: {ids[:5]}...)")
        
        # Verificar campos de data
        campos_data = [h for h in headers if h and any(palavra in str(h).lower() for palavra in ['data', 'vencimento', 'inicio', 'fim'])]
        if campos_data:
            print(f"  📅 Campos de data identificados: {', '.join(campos_data)}")
        
        # Verificar campos monetários
        campos_valor = [h for h in headers if h and any(palavra in str(h).lower() for palavra in ['valor', 'aluguel', 'preco', 'r$'])]
        if campos_valor:
            print(f"  💰 Campos monetários identificados: {', '.join(campos_valor)}")
        
        analise_completa[sheet_name] = {
            'dimensoes': {'linhas': max_row, 'colunas': max_col},
            'headers': headers,
            'colunas': analise_colunas,
            'comentarios': comentarios_encontrados,
            'validacoes': validacoes_encontradas
        }
    
    # Salvar análise em JSON
    print("\n" + "="*80)
    print("💾 SALVANDO ANÁLISE DETALHADA")
    print("="*80)
    
    with open('analise_planilha.json', 'w', encoding='utf-8') as f:
        json.dump(analise_completa, f, indent=2, ensure_ascii=False, default=str)
    
    print("\n✓ Análise salva em: analise_planilha.json")
    
    # Gerar mapeamento de campos
    print("\n" + "="*80)
    print("🗺️  MAPEAMENTO SUGERIDO PARA MIGRAÇÃO")
    print("="*80)
    
    for sheet_name, dados in analise_completa.items():
        if sheet_name.upper() in ['DASHBOARD', 'CONFIGURAÇÕES', 'CONFIG']:
            continue
        
        print(f"\n📋 Aba: {sheet_name}")
        for header in dados['headers']:
            if header:
                print(f"  '{header}' → campo do banco")
    
    wb.close()
    
    return analise_completa


if __name__ == "__main__":
    import sys
    
    caminho = 'ImobiPro.xlsx'
    if len(sys.argv) > 1:
        caminho = sys.argv[1]
    
    try:
        analise = analisar_planilha_completa(caminho)
        print("\n✓ Análise completa!")
        print("\nUse o arquivo 'analise_planilha.json' para ver todos os detalhes.")
    except FileNotFoundError:
        print(f"\n✗ Arquivo não encontrado: {caminho}")
        print("Use: python3 analisar_planilha_detalhado.py ImobiPro.xlsx")
    except Exception as e:
        print(f"\n✗ Erro ao analisar: {e}")
        import traceback
        traceback.print_exc()