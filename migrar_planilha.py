"""
================================================================================
IMOBIPRO - SCRIPT DE MIGRAÇÃO (VERSÃO CORRIGIDA)
================================================================================
Data: Janeiro 2026
Descrição: Migração corrigida baseada na análise detalhada da planilha.
           Usa os nomes EXATOS dos campos e implementa regras de negócio.
================================================================================
"""

import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from database.db_manager import DatabaseManager

class MigradorPlanilhaCorrigido:
    """
    Migrador corrigido com mapeamento exato dos campos da planilha.
    """
    
    def __init__(self, caminho_excel: str):
        self.caminho_excel = caminho_excel
        self.db = DatabaseManager()
        self.log_erros = []
        self.log_sucessos = []
        self.mapeamento_imoveis = {}  # ID_Propriedade (planilha) -> id (banco)
        self.mapeamento_pessoas = {}  # ID_Pessoa (planilha) -> id (banco)
        self.mapeamento_contratos = {}  # ID_Contrato (planilha) -> id (banco)
        
    def validar_arquivo(self) -> bool:
        if not os.path.exists(self.caminho_excel):
            print(f"✗ Arquivo não encontrado: {self.caminho_excel}")
            return False
        
        try:
            wb = load_workbook(self.caminho_excel, data_only=True)
            wb.close()
            print(f"✓ Arquivo Excel validado: {self.caminho_excel}")
            return True
        except Exception as e:
            print(f"✗ Erro ao abrir arquivo Excel: {e}")
            return False
    
    def converter_data(self, valor) -> str:
        """Converte valores de data para formato ISO."""
        if valor is None or valor == "":
            return None
        
        if isinstance(valor, datetime):
            return valor.strftime('%Y-%m-%d')
        
        if isinstance(valor, str):
            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                try:
                    dt = datetime.strptime(valor.strip(), fmt)
                    return dt.strftime('%Y-%m-%d')
                except:
                    continue
        
        return None
    
    def limpar_valor(self, valor) -> any:
        """Limpa e prepara valor para inserção."""
        if valor is None or valor == "":
            return None
        if isinstance(valor, str):
            return valor.strip()
        return valor
    
    def migrar_imoveis(self, wb):
        """Migra dados da aba 'imoveis' com campos corretos."""
        print("\n" + "="*80)
        print("MIGRANDO IMÓVEIS")
        print("="*80)
        
        try:
            ws = wb['imoveis']
        except KeyError:
            print("✗ Aba 'imoveis' não encontrada")
            return
        
        # Ler cabeçalhos
        headers = [cell.value for cell in ws[1]]
        print(f"Campos encontrados: {len([h for h in headers if h])} colunas")
        
        sucessos = 0
        erros = 0
        
        # Processar cada linha
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                # Criar dicionário com mapeamento exato
                dados_planilha = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        dados_planilha[header] = self.limpar_valor(row[i])
                
                # ID da planilha (para mapeamento)
                id_planilha = dados_planilha.get('ID_Propriedade')
                
                # Mapear para campos do banco
                dados_banco = {
                    'endereco_completo': dados_planilha.get('EnderecoCompleto'),
                    'inscricao_imobiliaria': dados_planilha.get('InscricaoImobiliaria'),
                    'tipo_imovel': dados_planilha.get('TipoImovel', 'Casa'),
                    'valor_iptu_anual': dados_planilha.get('ValorIPTUAnual'),
                    'forma_pagamento_iptu': dados_planilha.get('FormaPagamentoIPTU', 'Anual'),
                    'aluguel_pretendido': dados_planilha.get('AluguelPretendido'),
                    'condominio_sugerido': dados_planilha.get('CondominioSugerido'),
                    'dia_venc_condominio': dados_planilha.get('DiaVencCondominio'),
                    'valor_mercado': dados_planilha.get('ValorMercado'),
                    'data_aquisicao': self.converter_data(dados_planilha.get('DataAquisicao')),
                    'numero_hidrometro': dados_planilha.get('NumeroHidrometro'),
                    'numero_relogio_energia': dados_planilha.get('NumeroRelogioEnergia'),
                    'cidade': dados_planilha.get('Cidade', 'Campo Grande'),
                    'estado': dados_planilha.get('Estado', 'MS'),
                    'cep': dados_planilha.get('CEP'),
                    'observacoes': dados_planilha.get('Observacoes'),
                }
                
                # Validar campo obrigatório
                if not dados_banco['endereco_completo']:
                    raise ValueError("EnderecoCompleto é obrigatório")
                
                # Inserir no banco
                imovel_id = self.db.insert('imoveis', dados_banco)
                
                if imovel_id:
                    # Armazenar mapeamento
                    if id_planilha:
                        self.mapeamento_imoveis[int(id_planilha)] = imovel_id
                    
                    sucessos += 1
                    self.log_sucessos.append(f"Imóvel linha {idx}")
                    print(f"  ✓ Linha {idx}: {dados_banco['endereco_completo'][:50]}... (ID: {imovel_id})")
                else:
                    erros += 1
                    
            except Exception as e:
                erros += 1
                msg = f"Linha {idx}: {str(e)}"
                self.log_erros.append(msg)
                print(f"  ✗ {msg}")
        
        print(f"\n✓ Imóveis migrados: {sucessos}")
        if erros > 0:
            print(f"✗ Erros: {erros}")
    
    def migrar_pessoas(self, wb):
        """Migra dados da aba 'pessoas' com campos corretos."""
        print("\n" + "="*80)
        print("MIGRANDO PESSOAS")
        print("="*80)
        
        try:
            ws = wb['pessoas']
        except KeyError:
            print("✗ Aba 'pessoas' não encontrada")
            return
        
        headers = [cell.value for cell in ws[1]]
        print(f"Campos encontrados: {len([h for h in headers if h])} colunas")
        
        sucessos = 0
        erros = 0
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                # Mapear dados
                dados_planilha = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        dados_planilha[header] = self.limpar_valor(row[i])
                
                id_planilha = dados_planilha.get('ID_Pessoa')
                
                # Mapear para banco (nomes exatos da planilha!)
                dados_banco = {
                    'situacao': dados_planilha.get('Situação', 'Inquilino'),
                    'nome_completo': dados_planilha.get('NomeCompleto'),
                    'cpf_cnpj': dados_planilha.get('CPFNJP'),
                    'telefone': dados_planilha.get('Telefone'),
                    'email': dados_planilha.get('Email'),
                    'endereco_completo': dados_planilha.get('EnderecoCompleto'),
                    'cidade': dados_planilha.get('Cidade'),
                    'patrimonio': dados_planilha.get('Patrimonio'),
                    'estado_civil': dados_planilha.get('EstadoCivil'),
                    'nome_conjuge': dados_planilha.get('NomeConjuge'),
                    'cpf_conjuge': dados_planilha.get('CPFConjuge'),
                }
                
                if not dados_banco['nome_completo']:
                    raise ValueError("NomeCompleto é obrigatório")
                
                pessoa_id = self.db.insert('pessoas', dados_banco)
                
                if pessoa_id:
                    if id_planilha:
                        self.mapeamento_pessoas[int(id_planilha)] = pessoa_id
                    
                    sucessos += 1
                    self.log_sucessos.append(f"Pessoa linha {idx}")
                    print(f"  ✓ Linha {idx}: {dados_banco['nome_completo'][:40]}... (ID: {pessoa_id})")
                else:
                    erros += 1
                    
            except Exception as e:
                erros += 1
                msg = f"Linha {idx}: {str(e)}"
                self.log_erros.append(msg)
                print(f"  ✗ {msg}")
        
        print(f"\n✓ Pessoas migradas: {sucessos}")
        if erros > 0:
            print(f"✗ Erros: {erros}")
    
    def migrar_contratos(self, wb):
        """Migra dados da aba 'contratos' com campos corretos."""
        print("\n" + "="*80)
        print("MIGRANDO CONTRATOS")
        print("="*80)
        
        try:
            ws = wb['contratos']
        except KeyError:
            print("✗ Aba 'contratos' não encontrada")
            return
        
        headers = [cell.value for cell in ws[1]]
        print(f"Campos encontrados: {len([h for h in headers if h])} colunas")
        
        sucessos = 0
        erros = 0
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                dados_planilha = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        # Ignorar fórmulas, pegar apenas valores
                        valor = row[i]
                        if not isinstance(valor, str) or not valor.startswith('='):
                            dados_planilha[header] = self.limpar_valor(valor)
                
                id_planilha = dados_planilha.get('ID_Contrato')
                
                # Converter IDs da planilha para IDs do banco
                id_propriedade_planilha = dados_planilha.get('ID_Propriedade')
                id_inquilino_planilha = dados_planilha.get('ID_Inquilino')
                id_fiador_planilha = dados_planilha.get('ID_Fiador')
                
                # Tentar converter para int (podem vir como string de fórmula)
                try:
                    if id_propriedade_planilha:
                        id_propriedade_planilha = int(float(str(id_propriedade_planilha)))
                except:
                    pass
                
                try:
                    if id_inquilino_planilha:
                        id_inquilino_planilha = int(float(str(id_inquilino_planilha)))
                except:
                    pass
                
                try:
                    if id_fiador_planilha and str(id_fiador_planilha).strip():
                        id_fiador_planilha = int(float(str(id_fiador_planilha)))
                except:
                    id_fiador_planilha = None
                
                # Buscar IDs correspondentes no banco
                id_imovel = self.mapeamento_imoveis.get(id_propriedade_planilha)
                id_inquilino = self.mapeamento_pessoas.get(id_inquilino_planilha)
                id_fiador = self.mapeamento_pessoas.get(id_fiador_planilha) if id_fiador_planilha else None
                
                if not id_imovel or not id_inquilino:
                    raise ValueError(f"IDs não encontrados: Imóvel={id_propriedade_planilha}, Inquilino={id_inquilino_planilha}")
                
                dados_banco = {
                    'id_imovel': id_imovel,
                    'id_inquilino': id_inquilino,
                    'id_fiador': id_fiador,
                    'garantia': dados_planilha.get('Garantia', 'nenhuma'),
                    'inicio_contrato': self.converter_data(dados_planilha.get('InicioContrato')),
                    'fim_contrato': self.converter_data(dados_planilha.get('FimContrato')),
                    'valor_aluguel': dados_planilha.get('ValorAluguel'),
                    'dia_vencimento': dados_planilha.get('DiaVenc', 10),
                    'status_contrato': dados_planilha.get('StatusContrato', 'Ativo'),
                    'observacoes': dados_planilha.get('Obs'),
                }
                
                # Converter dia_vencimento para int
                try:
                    if dados_banco['dia_vencimento']:
                        dados_banco['dia_vencimento'] = int(float(str(dados_banco['dia_vencimento'])))
                except:
                    dados_banco['dia_vencimento'] = 10
                
                contrato_id = self.db.insert('contratos', dados_banco)
                
                if contrato_id:
                    if id_planilha:
                        self.mapeamento_contratos[int(id_planilha)] = contrato_id
                    
                    sucessos += 1
                    print(f"  ✓ Linha {idx}: Contrato ID {contrato_id} (Imóvel {id_imovel}, Status: {dados_banco['status_contrato']})")
                else:
                    erros += 1
                    
            except Exception as e:
                erros += 1
                msg = f"Linha {idx}: {str(e)}"
                self.log_erros.append(msg)
                print(f"  ✗ {msg}")
        
        print(f"\n✓ Contratos migrados: {sucessos}")
        if erros > 0:
            print(f"✗ Erros: {erros}")
    
    def migrar_despesas(self, wb):
        """Migra dados da aba 'despesas' com campos corretos."""
        print("\n" + "="*80)
        print("MIGRANDO DESPESAS")
        print("="*80)
        
        try:
            ws = wb['despesas']
        except KeyError:
            print("✗ Aba 'despesas' não encontrada")
            return
        
        headers = [cell.value for cell in ws[1]]
        sucessos = 0
        erros = 0
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                dados_planilha = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        valor = row[i]
                        if not isinstance(valor, str) or not valor.startswith('='):
                            dados_planilha[header] = self.limpar_valor(valor)
                
                # Converter ID_Propriedade
                id_propriedade_planilha = dados_planilha.get('Id_Propriedade')
                try:
                    if id_propriedade_planilha:
                        id_propriedade_planilha = int(float(str(id_propriedade_planilha)))
                except:
                    pass
                
                id_imovel = self.mapeamento_imoveis.get(id_propriedade_planilha)
                if not id_imovel:
                    raise ValueError(f"Imóvel não encontrado: {id_propriedade_planilha}")
                
                dados_banco = {
                    'id_imovel': id_imovel,
                    'mes_referencia': self.converter_data(dados_planilha.get('MesReferencia')),
                    'tipo_despesa': dados_planilha.get('TipoDespesa'),
                    'motivo_despesa': dados_planilha.get('MotivoDespesa'),
                    'valor_previsto': dados_planilha.get('ValorPrevisto'),
                    'valor_pago': dados_planilha.get('ValorPago'),
                    'vencimento_previsto': self.converter_data(dados_planilha.get('VencimentoPrevisto')),
                    'data_pagamento': self.converter_data(dados_planilha.get('DataPagamento')),
                    'observacoes': dados_planilha.get('Observacoes'),
                    'recorrente': 1 if dados_planilha.get('TipoDespesa') in ['Condomínio', 'IPTU'] else 0,
                }
                
                despesa_id = self.db.insert('despesas', dados_banco)
                if despesa_id:
                    sucessos += 1
                    print(f"  ✓ Linha {idx}: Despesa ID {despesa_id}")
                else:
                    erros += 1
                    
            except Exception as e:
                erros += 1
                msg = f"Linha {idx}: {str(e)}"
                self.log_erros.append(msg)
                print(f"  ✗ {msg}")
        
        print(f"\n✓ Despesas migradas: {sucessos}")
        if erros > 0:
            print(f"✗ Erros: {erros}")
    
    def migrar_receitas(self, wb):
        """Migra dados da aba 'receitas' com campos corretos."""
        print("\n" + "="*80)
        print("MIGRANDO RECEITAS")
        print("="*80)
        
        try:
            ws = wb['receitas']
        except KeyError:
            print("✗ Aba 'receitas' não encontrada")
            return
        
        headers = [cell.value for cell in ws[1]]
        sucessos = 0
        erros = 0
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                dados_planilha = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        valor = row[i]
                        if not isinstance(valor, str) or not valor.startswith('='):
                            dados_planilha[header] = self.limpar_valor(valor)
                
                id_contrato_planilha = dados_planilha.get('ID_Contrato')
                try:
                    if id_contrato_planilha:
                        id_contrato_planilha = int(id_contrato_planilha)
                except:
                    pass
                
                id_contrato = self.mapeamento_contratos.get(id_contrato_planilha)
                if not id_contrato:
                    raise ValueError(f"Contrato não encontrado: {id_contrato_planilha}")
                
                aluguel = dados_planilha.get('AluguelDevido', 0)
                condominio = dados_planilha.get('CondominioDevido', 0)
                iptu = dados_planilha.get('IPTUDevido', 0)
                desconto_multa = dados_planilha.get('Desconto(-) Multa (+)', 0)
                
                # Converter para float
                try:
                    aluguel = float(aluguel) if aluguel else 0
                    condominio = float(condominio) if condominio else 0
                    iptu = float(iptu) if iptu else 0
                    desconto_multa = float(desconto_multa) if desconto_multa else 0
                except:
                    pass
                
                dados_banco = {
                    'id_contrato': id_contrato,
                    'mes_referencia': self.converter_data(dados_planilha.get('MesReferencia')),
                    'aluguel_devido': aluguel,
                    'condominio_devido': condominio,
                    'iptu_devido': iptu,
                    'desconto_multa': desconto_multa,
                    'valor_total_devido': aluguel + condominio + iptu + desconto_multa,
                    'vencimento_previsto': self.converter_data(dados_planilha.get('VencimentoPrevisto')),
                    'data_recebimento': self.converter_data(dados_planilha.get('DataRecebimento')),
                    'valor_recebido': dados_planilha.get('ValorRecebido'),
                    'status': dados_planilha.get('Status', 'Pendente'),
                    'observacoes': dados_planilha.get('Observacoes'),
                }
                
                receita_id = self.db.insert('receitas', dados_banco)
                if receita_id:
                    sucessos += 1
                    print(f"  ✓ Linha {idx}: Receita ID {receita_id}")
                else:
                    erros += 1
                    
            except Exception as e:
                erros += 1
                msg = f"Linha {idx}: {str(e)}"
                self.log_erros.append(msg)
                print(f"  ✗ {msg}")
        
        print(f"\n✓ Receitas migradas: {sucessos}")
        if erros > 0:
            print(f"✗ Erros: {erros}")
    
    def migrar(self):
        """Executa migração completa."""
        print("\n" + "="*80)
        print("INICIANDO MIGRAÇÃO DA PLANILHA (VERSÃO CORRIGIDA)")
        print("="*80)
        print(f"Arquivo: {self.caminho_excel}")
        print(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        
        if not self.validar_arquivo():
            return False
        
        print("\nInicializando banco de dados...")
        self.db.initialize_database()
        
        print("\nCarregando planilha Excel...")
        wb = load_workbook(self.caminho_excel, data_only=True)
        
        # Migrar em ordem (respeitando dependências)
        self.migrar_imoveis(wb)
        self.migrar_pessoas(wb)
        self.migrar_contratos(wb)
        self.migrar_despesas(wb)
        self.migrar_receitas(wb)
        
        wb.close()
        
        # Relatório final
        print("\n" + "="*80)
        print("MIGRAÇÃO CONCLUÍDA")
        print("="*80)
        print(f"✓ Registros migrados: {len(self.log_sucessos)}")
        print(f"✗ Erros: {len(self.log_erros)}")
        
        if self.log_erros:
            print("\nPrimeiros 10 erros:")
            for erro in self.log_erros[:10]:
                print(f"  - {erro}")
        
        # Estatísticas
        print("\nEstatísticas do banco:")
        stats = self.db.get_estatisticas_dashboard()
        print(f"  Total de imóveis: {stats['total_imoveis']}")
        print(f"  Imóveis ocupados: {stats['imoveis_ocupados']}")
        print(f"  Contratos ativos: {stats['contratos_ativos']}")
        print(f"  Taxa de ocupação: {stats['taxa_ocupacao']:.1f}%")
        
        return True


if __name__ == "__main__":
    caminho = 'ImobiPro.xlsx'
    if len(sys.argv) > 1:
        caminho = sys.argv[1]
    
    migrador = MigradorPlanilhaCorrigido(caminho)
    sucesso = migrador.migrar()
    
    sys.exit(0 if sucesso else 1)